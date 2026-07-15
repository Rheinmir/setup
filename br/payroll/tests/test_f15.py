"""f15 UI — 3 màn hình + BẤT BIẾN THỊ GIÁC máy-kiểm (BR C15.1, C15.2, C15.3 → br/DESIGN.md).

Luật thị giác ở đây là ASSERT CỨNG, không phải lời dặn trong prompt.
Tương phản tính bằng SỐ (WCAG AAA ≥ 7:1), cấm ước lượng bằng mắt.
"""
import re
import threading
import unittest
import urllib.error
import urllib.request
from app import ui

FONT_CAM = ("Inter", "Roboto", "Arial", "Open Sans", "Helvetica")
FONT_HOP_LE = ("Plus Jakarta Sans", "Geist", "Clash Display", "PP Editorial New")


class _NoRedirect(urllib.request.HTTPErrorProcessor):
    """Đừng biến 3xx/4xx thành exception hay tự follow — trả thẳng response
    gốc để đọc .status/.headers (urlopen() mặc định tự follow POST 302 và
    NUỐT MẤT Set-Cookie của response gốc — phải chặn ở đây để lấy cookie)."""
    def http_response(self, request, response):
        return response
    https_response = http_response


def _lum(hexstr):
    """Độ chói tương đối WCAG."""
    h = hexstr.lstrip("#")
    if len(h) == 3:
        h = "".join(c * 2 for c in h)
    out = []
    for i in (0, 2, 4):
        c = int(h[i:i + 2], 16) / 255
        out.append(c / 12.92 if c <= 0.03928 else ((c + 0.055) / 1.055) ** 2.4)
    return 0.2126 * out[0] + 0.7152 * out[1] + 0.0722 * out[2]


def contrast(a, b):
    la, lb = _lum(a), _lum(b)
    hi, lo = max(la, lb), min(la, lb)
    return (hi + 0.05) / (lo + 0.05)


def token(css, block, name):
    """Lấy giá trị token màu trong một khối CSS (vd :root hoặc [data-theme=dark])."""
    m = re.search(re.escape(block) + r"[^{]*\{([^}]*)\}", css, re.S)
    if not m:
        return None
    m2 = re.search(rf"{re.escape(name)}\s*:\s*(#[0-9a-fA-F]{{3,8}})", m.group(1))
    return m2.group(1) if m2 else None


def token_any(css, blocks, name):
    for b in blocks:
        v = token(css, b, name)
        if v:
            return v
    return None


DARK_BLOCKS = ("[data-theme=dark]", '[data-theme="dark"]')


class TestUI(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.srv = ui.build_server(0)
        cls.port = cls.srv.server_address[1]
        threading.Thread(target=cls.srv.serve_forever, daemon=True).start()
        # C2.1 (cập nhật 2026-07-15): mọi route đòi session — tự "đăng nhập"
        # 1 lần cho cả lớp test này, KHÔNG kiểm gì (đúng cổng UX, không phải auth thật)
        # PHẢI dùng _NoRedirect — urlopen() mặc định tự follow POST 302 và
        # nuốt mất Set-Cookie của response gốc trước khi ta đọc được nó
        opener = urllib.request.build_opener(_NoRedirect)
        req = urllib.request.Request(f"http://127.0.0.1:{cls.port}/login", method="POST")
        r = opener.open(req, timeout=10)
        assert r.status == 302, "POST /login phải redirect 302"
        cls.cookie = r.headers.get("Set-Cookie", "").split(";")[0]
        assert cls.cookie.startswith("session="), f"thiếu cookie session, got {cls.cookie!r}"
        cls.html = cls.fetch("/")

    @classmethod
    def tearDownClass(cls):
        cls.srv.shutdown()

    @classmethod
    def fetch(cls, path):
        req = urllib.request.Request(f"http://127.0.0.1:{cls.port}{path}",
                                     headers={"Cookie": getattr(cls, "cookie", "")})
        with urllib.request.urlopen(req, timeout=10) as r:
            assert r.status == 200, path
            return r.read().decode()

    # ── Nội dung: engine phải hiện đúng số thật ────────────────────────────
    def test_bang_luong_hien_dung_luong_thuc_nhan_cua_ground_truth(self):
        self.assertIn("189.930.161", self.html)

    def test_phieu_luong_dung_layout_that(self):
        h = self.fetch("/payslip/GT-ROW9")
        self.assertIn("189.930.161", h)
        self.assertIn("21", h)
        for muc in ("Lương thực nhận", "Thuế TNCN"):
            self.assertIn(muc, h)

    def test_phieu_luong_hien_cot_rieng_phu_cap_truy_thu(self):
        # C8.8/FE-06 — PC_TRUY_THU phải là CỘT RIÊNG, luôn hiện (kể cả 0, GT-ROW9
        # không có ca truy thu) — không được gộp/ẩn vào ADJ_PLUS/ADJ_MINUS
        h = self.fetch("/payslip/GT-ROW9")
        self.assertIn("Phụ cấp truy thu/truy lĩnh", h)
        self.assertIn('href="/trace/GT-ROW9/PC_TRUY_THU"', h)

    def test_trace_cong_thuc_di_nguoc_duoc(self):
        h = self.fetch("/trace/GT-ROW9/NET_PAY")
        self.assertIn("NET_PAY", h)
        self.assertIn("GROSS", h)
        self.assertIn("C13", h)

    def test_trace_hien_tham_so_da_dung(self):
        self.assertIn("46.800.000", self.fetch("/trace/GT-ROW9/SI_EMP"))

    # ── §3.1 Absolute Zero: chống mẫu ──────────────────────────────────────
    def test_khong_dung_font_cam(self):
        for f in FONT_CAM:
            self.assertNotIn(f, self.html, f"font cấm '{f}' (design-template §3.1)")

    def test_co_font_hop_le(self):
        self.assertTrue(any(f in self.html for f in FONT_HOP_LE),
                        f"phải dùng một trong {FONT_HOP_LE}")

    def test_khong_co_vien_1px_solid(self):
        self.assertIsNone(re.search(r"border\s*:\s*1px\s+solid", self.html),
                          "viền 1px solid bị cấm — chiều sâu do BÓNG, không do đường kẻ (§3.1, §3.6)")

    # ── §3.6 Neumorphism recipe ────────────────────────────────────────────
    def test_cap_bong_doi_xung_sang_va_toi(self):
        self.assertIn("--sh-light", self.html)
        self.assertIn("--sh-dark", self.html)
        noi = [s for s in re.findall(r"box-shadow\s*:([^;]+);", self.html) if "inset" not in s]
        self.assertTrue(noi, "không có khối nổi nào")
        for s in noi:
            self.assertIn("--sh-light", s, "khối nổi thiếu bóng SÁNG (cặp bóng phải đủ 2)")
            self.assertIn("--sh-dark", s, "khối nổi thiếu bóng TỐI (cặp bóng phải đủ 2)")

    def test_co_trang_thai_lom_inset(self):
        self.assertIn("inset", self.html,
                      "phải có trạng thái lõm (inset) cho input/hàng chọn/nút nhấn — §3.6")

    def test_bo_goc_du_de_bong_doc_duoc(self):
        radii = [int(x) for x in re.findall(r"border-radius\s*:\s*(\d+)px", self.html)]
        self.assertTrue(radii, "không khai border-radius")
        self.assertGreaterEqual(max(radii), 12, "neumorphism cần bo tròn ≥12px — §3.6")

    def test_card_khong_co_mau_nen_rieng_khac_bg(self):
        self.assertNotIn("--card", self.html, "neumorphism KHÔNG có --card riêng — §4")
        self.assertIsNone(re.search(r"background(-color)?\s*:\s*(#fff\b|#ffffff\b|white\b)",
                                    self.html, re.I),
                          "card phải dùng CHUNG màu nền, không có nền trắng riêng — §3.6")

    def test_accent_lat_theo_theme(self):
        light = token(self.html, ":root", "--accent")
        dark = token_any(self.html, DARK_BLOCKS, "--accent")
        self.assertIsNotNone(light, "thiếu --accent ở theme sáng")
        self.assertIsNotNone(dark, "thiếu --accent ở theme tối")
        self.assertNotEqual(light.lower(), dark.lower(),
                            "accent phải LẬT theo theme (mặt sáng → accent tối; mặt tối → accent sáng) — §3.6")

    def test_accent_la_near_black_khong_phai_mau_ruc(self):
        # §3.6: "đối nghịch" là đối nghịch ĐỘ SÁNG, không phải thêm hue mới.
        # Accent cùng họ màu với nền (xanh trên nền xanh-xám) → CHÌM, dù pass AA.
        # Máy kiểm: độ bão hoà thấp (họ trung tính) + độ sáng ở cực trị.
        for ten, hexv, phai_toi in (("sáng", token(self.html, ":root", "--accent"), True),
                                    ("tối", token_any(self.html, DARK_BLOCKS, "--accent"), False)):
            h = hexv.lstrip("#")
            r, g, b = (int(h[i:i + 2], 16) / 255 for i in (0, 2, 4))
            mx, mn = max(r, g, b), min(r, g, b)
            l = (mx + mn) / 2
            s = 0 if mx == mn else (mx - mn) / (2 - mx - mn if l > 0.5 else mx + mn)
            self.assertLessEqual(
                round(s, 2), 0.20,
                f"accent theme {ten} = {hexv}: bão hoà {s:.2f} — đây là MÀU RỰC. "
                "Accent phải thuộc họ trung tính quanh đen/trắng (§3.6). "
                "Màu rực chỉ dùng cho semantic (đỏ = destructive).")
            if phai_toi:
                self.assertLessEqual(round(l, 2), 0.25,
                                     f"mặt sáng → accent phải TỐI (near-black), got {hexv}")
            else:
                self.assertGreaterEqual(round(l, 2), 0.75,
                                        f"mặt tối → accent phải SÁNG (near-white), got {hexv}")

    # ── A11y: tương phản là CON SỐ, cấm ước lượng bằng mắt ─────────────────
    def test_tuong_phan_chu_dat_AAA_o_ca_hai_theme(self):
        cases = [("sáng", token(self.html, ":root", "--bg"),
                  token(self.html, ":root", "--ink"), token(self.html, ":root", "--muted")),
                 ("tối", token_any(self.html, DARK_BLOCKS, "--bg"),
                  token_any(self.html, DARK_BLOCKS, "--ink"),
                  token_any(self.html, DARK_BLOCKS, "--muted"))]
        for ten, bg, ink, muted in cases:
            self.assertIsNotNone(bg, f"theme {ten}: thiếu --bg")
            for name, fg in (("--ink", ink), ("--muted", muted)):
                self.assertIsNotNone(fg, f"theme {ten}: thiếu {name}")
                r = contrast(fg, bg)
                self.assertGreaterEqual(
                    round(r, 2), 7.0,
                    f"theme {ten}: {name} {fg} trên {bg} = {r:.2f}:1 — chưa đạt AAA 7:1. "
                    "Pass ở sàn AA vẫn ra giao diện 'chìm' (design-template §3.6)")

    # ── Theme & số ─────────────────────────────────────────────────────────
    def test_co_toggle_sang_toi_va_chong_FOUC(self):
        self.assertIn("localStorage", self.html)
        self.assertIn("prefers-color-scheme", self.html)
        head = self.html.split("</head>")[0]
        self.assertIn("localStorage", head, "script chống FOUC phải nằm TRONG <head>")

    def test_tien_can_phai_va_dinh_dang_viet_nam(self):
        self.assertIn("tabular-nums", self.html)

    # ── C15.4 mass-upload — MỘT tab tối giản, không form nhập tay ──────────
    @classmethod
    def post(cls, path, body: bytes):
        req = urllib.request.Request(
            f"http://127.0.0.1:{cls.port}{path}", data=body, method="POST",
            headers={"Content-Type": "application/octet-stream",
                    "Cookie": getattr(cls, "cookie", "")})
        with urllib.request.urlopen(req, timeout=10) as r:
            assert r.status == 200, path
            return r.read().decode()

    def test_upload_form_dung_1_tab_khong_nhap_tay_tung_field(self):
        h = self.fetch("/upload")
        self.assertEqual(h.count('type="file"'), 1, "chỉ đúng 1 ô chọn file")
        # performed_by/reason là METADATA cho audit log (C14.2/FE-17), không phải
        # field lương/phụ cấp/chấm công — không phá nguyên tắc "1 tab, không nhập tay"
        inputs = re.findall(r'<input[^>]*name="([^"]+)"', h)
        self.assertEqual(sorted(inputs), ["performed_by", "period", "reason", "xlsx"],
                          f"/upload chỉ được có 4 input này, thấy {inputs}")

    def test_upload_post_ghi_va_xac_nhan_so_nhan_su(self):
        import io
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["employee_id", "ho_ten", "BASIC_SAL"])
        ws.append(["UP-101", "Nguyễn Test", 15000000])
        ws.append(["UP-102", "Trần Test", 16000000])
        buf = io.BytesIO()
        wb.save(buf)

        h = self.post("/upload?period=2099-03&performed_by=Test+Tester&reason=Unit+test",
                       buf.getvalue())
        self.assertIn("2", h)  # 2 nhân sự đã nạp
        self.assertIn('href="/"', h)

        from app import adapters
        rows = adapters.fetch_employees("2099-03")
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["employee_id"], "UP-101")
        self.assertEqual(rows[0]["BASIC_SAL"], 15000000)
        # ground-truth 2026-03 không bị đụng
        self.assertIn("189.930.161", self.fetch("/"))

    def test_upload_post_thieu_nguoi_hoac_ly_do_bi_tu_choi(self):
        # BR C14.2 — người thực hiện + lý do BẮT BUỘC, thiếu là từ chối (400), không âm thầm bỏ qua
        req = urllib.request.Request(
            f"http://127.0.0.1:{self.port}/upload?period=2099-04",
            data=b"khong quan trong", method="POST",
            headers={"Content-Type": "application/octet-stream", "Cookie": self.cookie})
        with self.assertRaises(urllib.error.HTTPError) as cm:
            urllib.request.urlopen(req, timeout=10)
        self.assertEqual(cm.exception.code, 400)

    def test_audit_screen_hien_dung_lan_upload_vua_lam(self):
        # tự làm 1 lần upload trong CHÍNH test này — không phụ thuộc thứ tự chạy
        # test khác (unittest chạy theo alphabet, "audit" < "upload")
        import io
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["employee_id"])
        ws.append(["AUDIT-CHK-1"])
        buf = io.BytesIO()
        wb.save(buf)
        self.post("/upload?period=2099-05&performed_by=Audit+Checker&reason=Kiem+tra+audit",
                  buf.getvalue())

        h = self.fetch("/audit")
        self.assertIn("Audit Checker", h)
        self.assertIn("Kiem tra audit", h)
        self.assertIn("2099-05", h)
        # chỉ-xem, có link lùi như 3 màn gốc
        self.assertIn('class="back"', h)

    # ── FE-23 Master Data — /params chỉ-xem cả 2 bộ tham số ─────────────────
    def test_params_screen_hien_ca_2_bo_tham_so(self):
        h = self.fetch("/params")
        self.assertIn("2020-01-01", h)
        self.assertIn("2026-01-01", h)
        self.assertIn('class="back"', h)

    def test_params_screen_hien_muc_chua_hieu_luc(self):
        # đúng chỗ tranh cãi 46,8tr/50,6tr trước đó — giờ hiện công khai, Ctrl+F ra
        h = self.fetch("/params")
        self.assertIn("ins_cap_bh_new", h)
        self.assertIn("50.600.000", h)

    # ── C15.6 / FE-31 — Phân bổ chi phí theo bộ phận (DỮ LIỆU DRAFT) ────────
    def test_bao_cao_phan_bo_theo_bo_phan_dung_tong(self):
        import io
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active
        cols = ["employee_id", "ho_ten", "phong_ban", "CONTRACT_TYPE", "NATIONALITY",
                "BASIC_SAL", "STD_DAYS", "OFFICIAL_DAYS", "DEPENDENT_CNT"]
        ws.append(cols)
        # VP301: 2 người · VP302: 1 người — draft, KHÔNG phải ground-truth
        ws.append(["D-01", "Draft A", "VP301", "Chính thức", "Việt Nam", 20_000_000, 22, 22, 0])
        ws.append(["D-02", "Draft B", "VP301", "Chính thức", "Việt Nam", 15_000_000, 22, 22, 1])
        ws.append(["D-03", "Draft C", "VP302", "Chính thức", "Việt Nam", 30_000_000, 22, 22, 2])
        buf = io.BytesIO(); wb.save(buf)
        self.post("/upload?period=2030-01&performed_by=Test+Draft&reason=Demo+FE-31+phan+bo",
                  buf.getvalue())

        h = self.fetch("/report/cost-by-dept?period=2030-01")
        self.assertIn("DỮ LIỆU DRAFT", h)   # C15.6 — phải cảnh báo rõ, không lẫn ground-truth
        self.assertIn("VP301", h)
        self.assertIn("VP302", h)
        self.assertIn("2", h)   # VP301 có 2 người
        self.assertIn('class="back"', h)

        # tổng toàn báo cáo phải bằng tổng TOTAL_CTY_COST cộng dồn từng người (đối chứng độc lập)
        from app import adapters, engine, params
        p = params.load("2030-01")
        tong = sum(engine.compute("TOTAL_CTY_COST", r, p)
                  for r in adapters.fetch_employees("2030-01"))
        self.assertIn(ui._fmt(tong), h)

    # ── C15.8 / FE-21 — Bảng công chi tiết, BÁO CÁO NHẠY CẢM chưa có auth ───
    def test_bang_cong_chi_tiet_canh_bao_bao_mat(self):
        # ground-truth thật (2026-03), không cần dữ liệu draft — chỉ đọc field ngày công có sẵn
        h = self.fetch("/report/attendance-detail?period=2026-03")
        self.assertIn("BÁO CÁO NHẠY CẢM", h)
        self.assertIn("CHƯA có kiểm soát truy cập", h)
        self.assertIn("GT-ROW9", h)
        self.assertIn('class="back"', h)
        # đúng ngày công chuẩn/tổng ngày hưởng thật của ground-truth (không hard-code)
        from app import adapters, engine, params
        p = params.load("2026-03")
        rec = adapters.fetch_employees("2026-03")[0]
        paid = engine.compute("PAID_DAYS", rec, p)
        self.assertIn(ui._fmt(paid), h)

    def test_bang_cong_chi_tiet_gate_qua_require_mask_money(self):
        # C2.1 — route đi qua auth.require(MASK_MONEY); vai lô đầu đủ quyền
        # nên vẫn 200, nhưng giả lập vai KHÔNG đủ quyền phải bị chặn 403
        import unittest.mock as mock
        from app import auth
        vai_gioi_han = {"name": "Test vai giới hạn", "perms": {auth.Perm.VIEW}}
        with mock.patch("app.auth.current_user", return_value=vai_gioi_han):
            req = urllib.request.Request(
                f"http://127.0.0.1:{self.port}/report/attendance-detail?period=2026-03",
                headers={"Cookie": self.cookie})
            with self.assertRaises(urllib.error.HTTPError) as cm:
                urllib.request.urlopen(req, timeout=10)
            self.assertEqual(cm.exception.code, 403)
        # vai lô đầu (mặc định, không patch) vẫn xem được bình thường
        req2 = urllib.request.Request(
            f"http://127.0.0.1:{self.port}/report/attendance-detail?period=2026-03",
            headers={"Cookie": self.cookie})
        self.assertEqual(urllib.request.urlopen(req2, timeout=10).status, 200)

    # ── C15.7 / FE-19 — Báo cáo Trình ký (Template 0), DỮ LIỆU DRAFT ────────
    def test_bao_cao_trinh_ky_gom_theo_du_an(self):
        import io
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active
        cols = ["employee_id", "ho_ten", "du_an", "CONTRACT_TYPE", "NATIONALITY",
                "BASIC_SAL", "STD_DAYS", "OFFICIAL_DAYS", "DEPENDENT_CNT"]
        ws.append(cols)
        ws.append(["S-01", "Sign A", "Dự án Bình Dương", "Chính thức", "Việt Nam",
                  18_000_000, 22, 22, 0])
        ws.append(["S-02", "Sign B", "Dự án Bình Dương", "Chính thức", "Việt Nam",
                  22_000_000, 22, 22, 1])
        ws.append(["S-03", "Sign C", "Dự án Long An", "Chính thức", "Việt Nam",
                  25_000_000, 22, 22, 0])
        buf = io.BytesIO(); wb.save(buf)
        self.post("/upload?period=2030-02&performed_by=Test+Draft&reason=Demo+FE-19+trinh+ky",
                  buf.getvalue())

        h = self.fetch("/report/signoff?period=2030-02")
        self.assertIn("DỮ LIỆU DRAFT", h)
        self.assertIn("Dự án Bình Dương", h)
        self.assertIn("Dự án Long An", h)
        self.assertIn("Sign A", h)
        self.assertIn("Sign C", h)
        self.assertIn('class="back"', h)
        # NET_PAY_HOME thật của từng người phải hiện đúng (đối chứng bằng engine, không hard-code)
        from app import adapters, engine, params
        p = params.load("2030-02")
        for r in adapters.fetch_employees("2030-02"):
            net = engine.compute("NET_PAY_HOME", r, p)
            self.assertIn(ui._fmt(net), h)

    # ── FE-20 Payroll Master — tải file phẳng CSV (không phải màn xem) ─────
    def test_export_payroll_master_tai_duoc_qua_route(self):
        import csv
        import io
        req = urllib.request.Request(
            f"http://127.0.0.1:{self.port}/export/payroll-master?period=2026-03",
            headers={"Cookie": self.cookie})
        with urllib.request.urlopen(req, timeout=10) as r:
            self.assertEqual(r.status, 200)
            self.assertIn("text/csv", r.headers.get("Content-Type", ""))
            self.assertIn("attachment", r.headers.get("Content-Disposition", ""))
            data = r.read().decode("utf-8")
        rows = list(csv.DictReader(io.StringIO(data)))
        self.assertEqual(rows[0]["employee_id"], "GT-ROW9")
        self.assertEqual(int(rows[0]["NET_PAY_HOME"]), 189_930_161)


class TestLoginGate(unittest.TestCase):
    """C2.1 (cập nhật 2026-07-15) — cổng session CHẶN THẬT mọi route, KHÔNG
    PHẢI xác thực bảo mật (login không kiểm mật khẩu/tài khoản nào)."""

    @classmethod
    def setUpClass(cls):
        cls.srv = ui.build_server(0)
        cls.port = cls.srv.server_address[1]
        threading.Thread(target=cls.srv.serve_forever, daemon=True).start()
        cls.opener = urllib.request.build_opener(_NoRedirect)

    @classmethod
    def tearDownClass(cls):
        cls.srv.shutdown()

    def _login(self):
        req = urllib.request.Request(f"http://127.0.0.1:{self.port}/login", method="POST")
        r = self.opener.open(req, timeout=10)
        return r.headers.get("Set-Cookie", "").split(";")[0]

    def test_vao_thang_trang_chu_chua_login_bi_redirect_ve_login(self):
        req = urllib.request.Request(f"http://127.0.0.1:{self.port}/")
        r = self.opener.open(req, timeout=10)
        self.assertEqual(r.status, 302)
        self.assertEqual(r.headers.get("Location"), "/login")

    def test_man_login_xem_duoc_khong_can_session(self):
        # /login CHÍNH NÓ không được đòi session, không thì kẹt vòng lặp redirect
        h = urllib.request.urlopen(f"http://127.0.0.1:{self.port}/login", timeout=10).read().decode()
        self.assertIn("Vào hệ thống", h)
        self.assertIn("HR C&amp;B", h)  # HTML đã escape & (đây là chữ trong thẻ login, không phải badge)
        # không có ô mật khẩu/tài khoản nào — đúng bản chất "cổng UX, không phải auth thật"
        self.assertNotIn('type="password"', h)

    def test_man_login_KHONG_hien_badge_khi_chua_dang_nhap(self):
        # bug đã sửa: _current_user là global, trang /login từng hiện badge vai
        # SÓT LẠI từ request trước — giờ phải phản ánh đúng session của request NÀY (rỗng)
        h = urllib.request.urlopen(f"http://127.0.0.1:{self.port}/login", timeout=10).read().decode()
        self.assertNotIn('class="who"', h)   # không badge vai
        self.assertNotIn("Đăng xuất", h)     # không link đăng xuất khi chưa đăng nhập

    def test_post_login_tao_session_cookie_va_redirect_ve_trang_chu(self):
        req = urllib.request.Request(f"http://127.0.0.1:{self.port}/login", method="POST")
        r = self.opener.open(req, timeout=10)
        self.assertEqual(r.status, 302)
        self.assertEqual(r.headers.get("Location"), "/")
        cookie = r.headers.get("Set-Cookie", "")
        self.assertIn("session=", cookie)

        # dùng cookie vừa nhận → vào trang chủ được thật
        req2 = urllib.request.Request(f"http://127.0.0.1:{self.port}/",
                                      headers={"Cookie": cookie.split(";")[0]})
        r2 = urllib.request.urlopen(req2, timeout=10)
        self.assertEqual(r2.status, 200)

    def test_logout_xoa_session_vao_lai_bi_chan_tiep(self):
        cookie = self._login()
        req_logout = urllib.request.Request(f"http://127.0.0.1:{self.port}/logout",
                                            headers={"Cookie": cookie})
        self.opener.open(req_logout, timeout=10)

        req3 = urllib.request.Request(f"http://127.0.0.1:{self.port}/", headers={"Cookie": cookie})
        r3 = self.opener.open(req3, timeout=10)
        self.assertEqual(r3.status, 302)
        self.assertEqual(r3.headers.get("Location"), "/login")

    def test_badge_hien_dung_vai_sau_khi_login(self):
        cookie = self._login()
        req2 = urllib.request.Request(f"http://127.0.0.1:{self.port}/", headers={"Cookie": cookie})
        h = urllib.request.urlopen(req2, timeout=10).read().decode()
        self.assertIn("HR C&amp;B", h)  # HTML đã escape &
        self.assertIn("Đăng xuất", h)


if __name__ == "__main__":
    unittest.main()
