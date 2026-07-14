"""upload.py — parse Excel mass-upload → list[dict] nhân sự (BR C15.4).

Header hàng 1 = đúng tên field `employees.json` (vd employee_id, BASIC_SAL...).
Không network, không parse công thức — chỉ đọc giá trị ô, giao adapters.py
ghi vào đúng chỗ fetch_employees đọc.
"""
from io import BytesIO

from openpyxl import load_workbook


def parse_employees_xlsx(content: bytes) -> list:
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() for h in next(rows) if h is not None]
    out = []
    for r in rows:
        if all(v is None for v in r):
            continue
        rec = {}
        for key, val in zip(header, r):
            rec[key] = 0 if val is None else val
        out.append(rec)
    return out
